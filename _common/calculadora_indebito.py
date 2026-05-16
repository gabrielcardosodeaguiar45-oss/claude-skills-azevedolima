"""Calculadora de indébito para ações de empréstimo não contratado / RMC / RCC.

Regime de atualização (conforme pedido nas iniciais do escritório):
  - Correção monetária: INPC (responsabilidade civil — STJ)
  - Juros de mora: 1% ao mês SIMPLES (juros legais, art. 406 CC c/c CTN —
    regra anterior à Lei 14.905/2024). Para casos posteriores a 30/08/2024,
    o procurador deve revisar se aplica SELIC-IPCA.
  - Dobro: art. 42, p. único, CDC

Geração de planilha Excel com:
  - 1 aba "RESUMO" com totais por contrato
  - 1 aba por contrato detalhando cada parcela descontada

Não usar para iniciais Bradesco AM (Patrick) — essas têm regime próprio na
skill inicial-bradesco.
"""
import os
import re
from datetime import date
from typing import List, Dict, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Patch E — Validador pré/pós-XLSX (2026-05-16)
# ----------------------------------------------------------------------------
# Caso paradigma VILSON/BANRISUL: XLSX gerado com R$ 50,00 × 29 meses fictícios
# (fallback do runner). Esta validação aborta a geração se algum contrato vier
# com valor_parcela <= 0 ou qtd_parcelas <= 0 — NÃO permite mais "calcular zero".
# ============================================================================

class CalculoIndebitoInvalidoError(RuntimeError):
    """Levantada quando dados de cálculo são insuficientes ou fictícios."""
    pass


def _validar_consistencia_competencia_fim(contratos: List[Dict]) -> List[str]:
    """Valida que `competencia_fim_str` é consistente com `situacao`.

    Aviso (não-fatal): se contrato está com `situacao=Excluído` ou `Encerrado`
    e NÃO tem `competencia_fim_str`/`data_exclusao_str`, o cálculo vai projetar
    descontos até o mês atual — pode somar parcelas inexistentes.

    Caso paradigma VILSON / BANRISUL: contrato portado em 26/08/2025 (excluído
    08/2025), mas o calculator projetou até 05/2026 — 10 parcelas inventadas.

    Returns: lista de avisos (vazia se OK).
    """
    avisos = []
    for i, c in enumerate(contratos, 1):
        numero = c.get('numero') or c.get('contrato') or '?'
        sit = (c.get('situacao') or '').strip().lower()
        comp_fim = (c.get('competencia_fim_str') or c.get('competencia_fim') or '').strip()
        data_exc = (c.get('data_exclusao_str') or c.get('data_exclusao') or '')

        if sit in ('excluído', 'excluido', 'encerrado'):
            if not comp_fim and not data_exc:
                avisos.append(
                    f'Contrato {i} (nº {numero}): situação "{sit.upper()}" mas '
                    f'sem `competencia_fim`/`data_exclusao` no dict. O cálculo '
                    f'vai projetar descontos até o mês atual e pode somar '
                    f'parcelas inexistentes. Preencher antes de gerar XLSX.'
                )
    return avisos


def _validar_contratos_para_calculo(contratos: List[Dict]) -> None:
    """Aborta se algum contrato a calcular está incompleto.

    Mesma filosofia do Patch C (validar_contratos_obrigatorios da skill
    inicial-nao-contratado), mas focado nos campos exigidos pelo cálculo:
    valor_parcela > 0, qtd_parcelas > 0, competencia_inicio.
    """
    erros = []
    for i, c in enumerate(contratos, 1):
        numero = c.get('numero') or c.get('contrato') or '?'
        prefix = f'Contrato {i} (nº {numero})'

        vp_raw = (c.get('valor_parcela_float') or c.get('valor_parcela')
                  or c.get('valor_parcela_str'))
        try:
            if isinstance(vp_raw, (int, float)):
                vp = float(vp_raw)
            elif isinstance(vp_raw, str):
                vp = float(re.sub(r'[^\d,.-]', '', vp_raw).replace('.', '').replace(',', '.') or '0')
            else:
                vp = 0.0
        except (ValueError, TypeError):
            vp = 0.0
        if vp <= 0:
            erros.append(f'{prefix}: valor_parcela inválido ({vp_raw!r}).')

        qtd = c.get('qtd_parcelas') or 0
        try:
            qtd = int(qtd)
        except (ValueError, TypeError):
            qtd = 0
        if qtd <= 0:
            erros.append(f'{prefix}: qtd_parcelas inválido ({c.get("qtd_parcelas")!r}).')

        ci = c.get('competencia_inicio_str') or c.get('competencia_inicio') or ''
        if not re.match(r'^\d{2}/\d{4}$', str(ci).strip()):
            erros.append(f'{prefix}: competência início inválida ({ci!r}).')

    if erros:
        raise CalculoIndebitoInvalidoError(
            'XLSX de indébito NÃO PODE ser gerado — dados insuficientes:\n'
            + '\n'.join(f'  • {e}' for e in erros)
            + '\n\nFallbacks fictícios (R$ 50,00, 84 parcelas, "01/2021") foram '
              'BANIDOS em 2026-05-16 (caso paradigma VILSON/BANRISUL).'
        )

# Import indices_oficiais — mesmo diretório
import sys
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)
from indices_oficiais import (inpc_acumulado_entre, corrigir_inpc,
                                juros_simples_mes, meses_entre,
                                INPC_ULTIMO_MES)


# ========================================
# Estilos
# ========================================
_BORDA = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999'),
)
_FILL_TITULO = PatternFill('solid', fgColor='305496')
_FILL_CAB = PatternFill('solid', fgColor='8EA9DB')
_FILL_TOTAL = PatternFill('solid', fgColor='FFE699')
_FILL_DOBRO = PatternFill('solid', fgColor='C6EFCE')
_FONT_TITULO = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
_FONT_CAB = Font(name='Calibri', size=11, bold=True)
_FONT_TOTAL = Font(name='Calibri', size=11, bold=True)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
_RIGHT = Alignment(horizontal='right', vertical='center')


# ========================================
# Helpers
# ========================================

def _parse_brl(valor_str: str) -> float:
    """'R$37,10' / '37,10' / '1.234,56' → float."""
    if isinstance(valor_str, (int, float)):
        return float(valor_str)
    if not valor_str:
        return 0.0
    s = str(valor_str).strip()
    s = re.sub(r'[Rr]\$\s*', '', s)
    s = s.replace('.', '').replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_competencia(comp_str: str) -> Optional[Tuple[int, int]]:
    """'08/2020' → (2020, 8). Retorna None se não parsear."""
    if not comp_str:
        return None
    m = re.match(r'(\d{1,2})/(\d{4})', str(comp_str).strip())
    if not m:
        return None
    mes, ano = int(m.group(1)), int(m.group(2))
    if 1 <= mes <= 12:
        return (ano, mes)
    return None


def _iter_meses(ini: Tuple[int, int], fim: Tuple[int, int]):
    a, m = ini
    while (a, m) <= fim:
        yield (a, m)
        m += 1
        if m > 12:
            m = 1
            a += 1


# ========================================
# Cálculo por contrato
# ========================================

def calcular_contrato(contrato: Dict, data_apuracao: Optional[date] = None,
                       taxa_juros_mes_pct: float = 1.0) -> Dict:
    """Calcula o indébito de um contrato.

    Args:
        contrato: dict com campos:
            - numero / contrato
            - banco / banco_nome
            - valor_parcela / valor_parcela_str (R$ XX,XX ou float)
            - qtd_parcelas (int)
            - competencia_inicio / competencia_inicio_str (MM/AAAA)
            - competencia_fim / competencia_fim_str (MM/AAAA, opcional)
            - situacao (Ativo / Excluído / Encerrado)
        data_apuracao: data do cálculo. Default = hoje.
        taxa_juros_mes_pct: % ao mês (default 1.0 = juros legais)

    Returns:
        dict {
            'contrato': str,
            'banco': str,
            'qtd_parcelas': int,
            'valor_parcela': float,
            'meses_pagos': int,
            'parcelas': [list de dicts mensais],
            'soma_pagos': float (sem correção),
            'soma_corrigida': float (INPC),
            'soma_juros': float,
            'total_simples': float (corrigido + juros),
            'total_dobrado': float (art. 42 CDC)
        }
    """
    if data_apuracao is None:
        data_apuracao = date.today()
    # Coerções
    numero = str(contrato.get('numero') or contrato.get('contrato') or '')
    banco = str(contrato.get('banco_nome') or contrato.get('banco') or '')
    vp = (contrato.get('valor_parcela_str') or contrato.get('valor_parcela')
          or contrato.get('valor_parcela_float') or 0)
    valor_parcela = _parse_brl(vp) if not isinstance(vp, (int, float)) else float(vp)
    qtd_parcelas = int(contrato.get('qtd_parcelas') or 0)
    comp_ini = (_parse_competencia(contrato.get('competencia_inicio_str')) or
                _parse_competencia(contrato.get('competencia_inicio')))
    comp_fim_extra = (_parse_competencia(contrato.get('competencia_fim_str')) or
                     _parse_competencia(contrato.get('competencia_fim')))
    situacao = (contrato.get('situacao') or '').strip()

    parcelas = []
    if not comp_ini or qtd_parcelas <= 0 or valor_parcela <= 0:
        return {
            'contrato': numero,
            'banco': banco,
            'qtd_parcelas': qtd_parcelas,
            'valor_parcela': valor_parcela,
            'meses_pagos': 0,
            'parcelas': [],
            'soma_pagos': 0,
            'soma_corrigida': 0,
            'soma_juros': 0,
            'total_simples': 0,
            'total_dobrado': 0,
            'alerta': 'Dados insuficientes (sem competência início ou qtd_parcelas ou valor_parcela)',
        }

    # Determinar último mês a contar
    # Regra:
    #   - Se há competencia_fim_str (encerrado): vai até ela
    #   - Senão (ativo): conta até qtd_parcelas OU mês de hoje, o que for primeiro
    a_ini, m_ini = comp_ini
    if comp_fim_extra:
        a_fim, m_fim = comp_fim_extra
    else:
        # Mês limite = comp_ini + qtd_parcelas - 1
        m_total = m_ini + qtd_parcelas - 1
        a_fim = a_ini + (m_total - 1) // 12
        m_fim = ((m_total - 1) % 12) + 1
    # Não passa do mês de apuração
    if (a_fim, m_fim) > (data_apuracao.year, data_apuracao.month):
        a_fim, m_fim = data_apuracao.year, data_apuracao.month

    mes_apuracao = (data_apuracao.year, data_apuracao.month)
    for (ano, mes) in _iter_meses((a_ini, m_ini), (a_fim, m_fim)):
        fator_inpc = inpc_acumulado_entre((ano, mes), mes_apuracao)
        n_meses = meses_entre(date(ano, mes, 1), data_apuracao)
        valor_corr = valor_parcela * fator_inpc
        juros = valor_parcela * (taxa_juros_mes_pct / 100.0) * n_meses
        total_simples = valor_corr + juros
        total_dobrado = total_simples * 2
        parcelas.append({
            'competencia': f'{mes:02d}/{ano}',
            'valor_original': valor_parcela,
            'fator_inpc': fator_inpc,
            'valor_corrigido': valor_corr,
            'meses_juros': n_meses,
            'juros': juros,
            'total_simples': total_simples,
            'total_dobrado': total_dobrado,
        })

    soma_pagos = sum(p['valor_original'] for p in parcelas)
    soma_corrigida = sum(p['valor_corrigido'] for p in parcelas)
    soma_juros = sum(p['juros'] for p in parcelas)
    total_simples = sum(p['total_simples'] for p in parcelas)
    total_dobrado = sum(p['total_dobrado'] for p in parcelas)
    return {
        'contrato': numero,
        'banco': banco,
        'qtd_parcelas': qtd_parcelas,
        'valor_parcela': valor_parcela,
        'meses_pagos': len(parcelas),
        'parcelas': parcelas,
        'soma_pagos': soma_pagos,
        'soma_corrigida': soma_corrigida,
        'soma_juros': soma_juros,
        'total_simples': total_simples,
        'total_dobrado': total_dobrado,
        'situacao': situacao,
    }


# ========================================
# Geração de Excel
# ========================================

def _set_brl(cell, valor):
    cell.value = valor
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = _RIGHT


def calcular_dano_moral(n_contratos: int) -> Dict:
    """Calcula o dano moral pleiteado conforme regra fixa do escritório:
      - 1 contrato                → R$ 15.000,00
      - 2+ contratos (cumulativo) → R$ 5.000,00 × N

    Returns: {'valor': float, 'criterio': str}
    """
    if n_contratos <= 0:
        return {'valor': 0.0, 'criterio': 'sem contratos'}
    if n_contratos == 1:
        return {'valor': 15000.0,
                'criterio': 'R$ 15.000,00 (1 contrato isolado)'}
    return {'valor': 5000.0 * n_contratos,
            'criterio': f'R$ 5.000,00 × {n_contratos} contratos'}


def calcular_valor_causa_nc(
    contratos: List[Dict],
    data_apuracao: Optional[date] = None,
    dano_moral_unico: float = 15000.0,
    dano_temporal: float = 5000.0,
) -> Dict:
    """Calcula o VALOR DA CAUSA para ação NC/RMC/RCC.

    Fórmula oficial do escritório (UNIFICADA inicial+XLSX, 2026-05-16):

        valor_causa = (TODAS_parcelas_descontadas × valor_parcela × 2)
                      + dano_moral
                      + dano_temporal

    **NÃO aplica prescrição retroativa.** Tese do trato sucessivo: em
    descontos mensais consecutivos, o termo inicial da prescrição flui
    do ÚLTIMO desconto, não do primeiro. Logo, ajuizando dentro de 5
    anos do último desconto, TODAS as parcelas históricas são impugnáveis
    (mesmo as anteriores a 5 anos do ajuizamento).

    Respeita:
      - `competencia_fim` quando o contrato foi extinto (portabilidade,
        quitação) antes do mês de apuração — não soma parcelas que não
        aconteceram
      - mês de apuração (não soma parcelas projetadas para o futuro)

    NÃO respeita prescrição quinquenal — pelas razões acima.

    Args:
        contratos: lista de dicts no formato `formatar_contrato_para_template`
        data_apuracao: default = hoje
        dano_moral_unico: R$ 15.000 para 1 contrato isolado
        dano_temporal: R$ 5.000 (REsp 1.737.412/SP — teoria do desvio produtivo)

    Returns:
        {
          'valor_causa': float,
          'soma_parcelas': float,         # TODAS as parcelas descontadas
          'dobro': float,
          'dano_moral': float,
          'dano_temporal': float,
          'qtd_parcelas_total': int,
          'detalhes_por_contrato': [...],
        }
    """
    if data_apuracao is None:
        data_apuracao = date.today()
    if not contratos:
        return {'valor_causa': 0.0, 'soma_parcelas': 0.0,
                'dobro': 0.0, 'dano_moral': 0.0, 'dano_temporal': 0.0,
                'qtd_parcelas_total': 0,
                'detalhes_por_contrato': []}

    soma_total = 0.0
    qtd_total = 0
    detalhes = []

    for c in contratos:
        # Valor da parcela
        vp = (c.get('valor_parcela_float') or c.get('valor_parcela'))
        if isinstance(vp, str):
            vp = _parse_brl(vp)
        vp = float(vp or 0)
        # Qtd parcelas
        qtd = int(c.get('qtd_parcelas') or 0)
        # Competência início
        ci = _parse_competencia(
            c.get('competencia_inicio_str') or c.get('competencia_inicio'))
        # Competência fim (HISCON ou default = parcelas projetadas)
        cf_str = c.get('competencia_fim_str') or c.get('competencia_fim')
        cf = _parse_competencia(cf_str) if cf_str else None

        if not ci or qtd <= 0 or vp <= 0:
            detalhes.append({
                'contrato': c.get('numero') or c.get('contrato'),
                'erro': 'dados insuficientes',
            })
            continue

        ano_ini, mes_ini = ci
        # Limite: min(competencia_fim, mês_apuração, ci+qtd-1)
        if cf:
            ano_lim, mes_lim = cf
        else:
            total_m = mes_ini + qtd - 1
            ano_lim = ano_ini + (total_m - 1) // 12
            mes_lim = ((total_m - 1) % 12) + 1
        if (ano_lim, mes_lim) > (data_apuracao.year, data_apuracao.month):
            ano_lim, mes_lim = data_apuracao.year, data_apuracao.month

        # Itera mês a mês — SOMA TODAS as parcelas (trato sucessivo)
        qtd_hist = 0
        soma_hist = 0.0
        a, m = ano_ini, mes_ini
        while (a, m) <= (ano_lim, mes_lim):
            qtd_hist += 1
            soma_hist += vp
            m += 1
            if m > 12:
                m = 1
                a += 1
        soma_total += soma_hist
        qtd_total += qtd_hist
        detalhes.append({
            'contrato': c.get('numero') or c.get('contrato'),
            'valor_parcela': vp,
            'qtd_parcelas': qtd_hist,
            'soma_parcelas': soma_hist,
            'dobro_parcial': soma_hist * 2,
        })

    dobro = soma_total * 2
    n_contratos = len(contratos)
    if n_contratos == 1:
        dm = dano_moral_unico
    else:
        dm = 5000.0 * n_contratos
    valor_causa = dobro + dm + dano_temporal
    return {
        'valor_causa': valor_causa,
        'soma_parcelas': soma_total,
        'dobro': dobro,
        'dano_moral': dm,
        'dano_temporal': dano_temporal,
        'qtd_parcelas_total': qtd_total,
        'detalhes_por_contrato': detalhes,
        'data_apuracao': data_apuracao.isoformat(),
    }


def gerar_excel_indebito(
    contratos: List[Dict],
    cliente_nome: str,
    output_path: str,
    data_apuracao: Optional[date] = None,
    taxa_juros_mes_pct: float = 1.0,
) -> str:
    """Gera planilha Excel com cálculo de indébito em ABA ÚNICA — pronta para
    exportar como PDF.

    Estrutura da aba (de cima para baixo):
      - Título com cliente e data de apuração
      - Para CADA contrato:
          • Cabeçalho do contrato (banco, número, situação, valor parcela)
          • Tabela mensal: competência | valor | fator INPC | corrigido |
                            meses juros | juros 1% a.m. | total simples |
                            total em dobro
          • Linha SUBTOTAL do contrato
      - SUBTOTAL GERAL (somatório de todos os contratos em dobro)
      - DANO MORAL (regra: 15k×1 ou 5k×N)
      - TOTAL GERAL DA AÇÃO

    Args:
        contratos: lista de dicts (formato extrator_hiscon ou kit-juridico)
        cliente_nome: nome para o cabeçalho
        output_path: caminho do .xlsx
        data_apuracao: default = hoje
        taxa_juros_mes_pct: 1.0 default

    Returns:
        output_path
    """
    if data_apuracao is None:
        data_apuracao = date.today()

    # Patch E (2026-05-16) — Validação pré-XLSX
    # Aborta se algum contrato vem com valor zero, qtd inválida, competência
    # vazia (sintomas de fallback fictício do tipo R$ 50,00 × 84 parcelas).
    _validar_contratos_para_calculo(contratos)

    # Patch G (2026-05-16) — Aviso de consistência competencia_fim vs situação
    # Não aborta (alguns templates aceitam alerta), mas registra para o caller
    # poder propagar como ALERTA na inicial.
    avisos_consistencia = _validar_consistencia_competencia_fim(contratos)
    if avisos_consistencia:
        # Imprime para o stderr (caller pode capturar) e segue
        import sys
        for a in avisos_consistencia:
            print(f'⚠ INCONSISTÊNCIA HISCON: {a}', file=sys.stderr)

    calculos = [calcular_contrato(c, data_apuracao, taxa_juros_mes_pct)
                for c in contratos]

    # Patch G (2026-05-16) — Validação pós-cálculo:
    # detecta se há parcelas projetadas APÓS a data_exclusao do contrato.
    # Sintoma: cálculo somando meses inexistentes (porque o contrato foi
    # extinto e a calculadora ignorou competencia_fim).
    for cidx, calc in enumerate(calculos):
        cont = contratos[cidx]
        data_exc_raw = (cont.get('data_exclusao') or cont.get('data_exclusao_str'))
        if not data_exc_raw:
            continue
        from datetime import datetime as _dt
        data_exc = None
        if isinstance(data_exc_raw, _dt):
            data_exc = data_exc_raw
        elif isinstance(data_exc_raw, str):
            for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
                try:
                    data_exc = _dt.strptime(data_exc_raw[:19], fmt)
                    break
                except ValueError:
                    continue
        if data_exc is None:
            continue
        # Se alguma parcela tem competência > mês da exclusão, é bug
        for p in calc.get('parcelas', []):
            comp = p.get('competencia', '')
            try:
                mes_s, ano_s = comp.split('/')
                p_dt = _dt(int(ano_s), int(mes_s), 1)
                # Permite até a competência da própria exclusão (mês fechado)
                if (p_dt.year, p_dt.month) > (data_exc.year, data_exc.month):
                    raise CalculoIndebitoInvalidoError(
                        f'Contrato nº {cont.get("numero")}: cálculo gerou '
                        f'parcela em {comp} mas o contrato foi EXCLUÍDO em '
                        f'{data_exc.strftime("%d/%m/%Y")}. Provável bug de '
                        f'projeção — verificar `competencia_fim_str` no '
                        f'contrato e ajustar antes de gerar XLSX.'
                    )
            except (ValueError, AttributeError):
                continue

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('RESUMO', 0)

    # Cabeçalhos da tabela mensal (8 colunas — referência única)
    cabs_mensal = ['Competência', 'Valor original', 'Fator INPC',
                    'Valor corrigido', 'Meses (juros)', 'Juros 1% a.m.',
                    'Total simples', 'Total em dobro (art. 42 CDC)']

    # ===== TÍTULO PRINCIPAL =====
    ws.merge_cells('A1:H1')
    ws['A1'] = f'CÁLCULO DE INDÉBITO — {cliente_nome.upper()}'
    ws['A1'].font = _FONT_TITULO
    ws['A1'].fill = _FILL_TITULO
    ws['A1'].alignment = _CENTER

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Data de apuração: {data_apuracao.strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(bold=True, italic=True)
    ws['A2'].alignment = _CENTER

    ws.merge_cells('A3:H3')
    ws['A3'] = (f'Regime: correção INPC + juros {taxa_juros_mes_pct:.1f}% a.m. '
                f'simples + dobro (art. 42 CDC). Último INPC disponível: '
                f'{INPC_ULTIMO_MES[1]:02d}/{INPC_ULTIMO_MES[0]}')
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = _CENTER

    ws.row_dimensions[1].height = 24

    row = 5  # 1 linha em branco entre cabeçalho e primeiro contrato

    # ===== UM BLOCO POR CONTRATO =====
    for idx, calc in enumerate(calculos):
        # Título do contrato
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1,
                 value=(f'CONTRATO Nº {calc["contrato"]} — {calc["banco"]}  '
                        f'(situação: {calc.get("situacao", "")}, valor parcela: '
                        f'R$ {calc["valor_parcela"]:,.2f}, '
                        f'{calc["meses_pagos"]} meses descontados)')
                 ).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=1).fill = _FILL_TITULO
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.row_dimensions[row].height = 22
        row += 1

        # Cabeçalho da tabela mensal
        for i, c in enumerate(cabs_mensal, 1):
            cell = ws.cell(row=row, column=i, value=c)
            cell.font = _FONT_CAB
            cell.fill = _FILL_CAB
            cell.alignment = _CENTER
            cell.border = _BORDA
        ws.row_dimensions[row].height = 30
        row += 1

        # Linhas mensais
        for p in calc['parcelas']:
            ws.cell(row=row, column=1, value=p['competencia']).alignment = _CENTER
            _set_brl(ws.cell(row=row, column=2), p['valor_original'])
            cf = ws.cell(row=row, column=3, value=p['fator_inpc'])
            cf.number_format = '0.000000'
            cf.alignment = _CENTER
            _set_brl(ws.cell(row=row, column=4), p['valor_corrigido'])
            ws.cell(row=row, column=5, value=p['meses_juros']).alignment = _CENTER
            _set_brl(ws.cell(row=row, column=6), p['juros'])
            _set_brl(ws.cell(row=row, column=7), p['total_simples'])
            _set_brl(ws.cell(row=row, column=8), p['total_dobrado'])
            ws.cell(row=row, column=8).fill = _FILL_DOBRO
            ws.cell(row=row, column=8).font = Font(bold=True)
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = _BORDA
            row += 1

        # Subtotal do contrato
        ws.cell(row=row, column=1, value=f'SUBTOTAL CONTRATO Nº {calc["contrato"]}').font = _FONT_TOTAL
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = _FILL_TOTAL
        _set_brl(ws.cell(row=row, column=2), calc['soma_pagos'])
        ws.cell(row=row, column=2).font = _FONT_TOTAL
        _set_brl(ws.cell(row=row, column=4), calc['soma_corrigida'])
        ws.cell(row=row, column=4).font = _FONT_TOTAL
        _set_brl(ws.cell(row=row, column=6), calc['soma_juros'])
        ws.cell(row=row, column=6).font = _FONT_TOTAL
        _set_brl(ws.cell(row=row, column=7), calc['total_simples'])
        ws.cell(row=row, column=7).font = _FONT_TOTAL
        _set_brl(ws.cell(row=row, column=8), calc['total_dobrado'])
        ws.cell(row=row, column=8).font = Font(bold=True, color='006100', size=12)
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = _BORDA
        row += 1

        # Linha em branco entre contratos
        row += 1

    # ===== SUBTOTAL GERAL =====
    soma_pagos_total = sum(c['soma_pagos'] for c in calculos)
    soma_simples_total = sum(c['total_simples'] for c in calculos)
    soma_dobrado_total = sum(c['total_dobrado'] for c in calculos)
    ws.cell(row=row, column=1,
             value=f'SUBTOTAL GERAL — {len(calculos)} contrato(s) em dobro'
             ).font = _FONT_TOTAL
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _FILL_TOTAL
    _set_brl(ws.cell(row=row, column=2), soma_pagos_total)
    ws.cell(row=row, column=2).font = _FONT_TOTAL
    _set_brl(ws.cell(row=row, column=7), soma_simples_total)
    ws.cell(row=row, column=7).font = _FONT_TOTAL
    _set_brl(ws.cell(row=row, column=8), soma_dobrado_total)
    ws.cell(row=row, column=8).font = Font(bold=True, color='006100', size=12)
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = _BORDA
    row += 1

    # ===== DANO MORAL =====
    n_contratos = len(calculos)
    dm = calcular_dano_moral(n_contratos)
    ws.cell(row=row, column=1,
             value='DANO MORAL (regra fixa do escritório)').font = _FONT_TOTAL
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _FILL_TOTAL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws.cell(row=row, column=2, value=dm['criterio']).font = Font(italic=True)
    ws.cell(row=row, column=2).alignment = _LEFT
    _set_brl(ws.cell(row=row, column=8), dm['valor'])
    ws.cell(row=row, column=8).font = Font(bold=True, color='006100', size=12)
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = _BORDA
    row += 1

    # ===== TOTAL GERAL =====
    total_geral = soma_dobrado_total + dm['valor']
    ws.cell(row=row, column=1, value='TOTAL GERAL DA AÇÃO').font = Font(
        name='Calibri', size=13, bold=True, color='FFFFFF')
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _FILL_TITULO
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws.cell(row=row, column=2,
             value='Subtotal em dobro + Dano moral').font = Font(
        italic=True, color='FFFFFF')
    ws.cell(row=row, column=2).alignment = _LEFT
    _set_brl(ws.cell(row=row, column=8), total_geral)
    ws.cell(row=row, column=8).font = Font(
        name='Calibri', size=14, bold=True, color='FFFF00')
    ws.row_dimensions[row].height = 28
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = _BORDA

    # Larguras (otimizadas para PDF A4 paisagem)
    for col, w in zip('ABCDEFGH', [13, 14, 11, 16, 11, 14, 16, 20]):
        ws.column_dimensions[col].width = w

    # Configurações de impressão (PDF)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = quantas páginas precisar na vertical
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path


# ========================================
# Leitura de Excel existente
# ========================================

def ler_total_geral_xlsx(path: str) -> Optional[Dict]:
    """Lê o 'TOTAL GERAL DA AÇÃO' de um Excel gerado por `gerar_excel_indebito`.

    Procura na aba 'RESUMO' a linha cujo texto da coluna A contém
    "TOTAL GERAL" e pega o valor da coluna H.

    Returns:
        dict com {
            'total_geral': float,
            'subtotal_dobrado': float,
            'dano_moral': float,
            'data_apuracao': str,  # do cabeçalho da planilha
        } ou None se não conseguir parsear.
    """
    from openpyxl import load_workbook
    if not os.path.exists(path):
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if 'RESUMO' not in wb.sheetnames:
            return None
        ws = wb['RESUMO']
        total_geral = None
        subtotal_dobrado = None
        dano_moral = None
        data_apuracao = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            primeira = row[0]
            if primeira is None:
                continue
            txt = str(primeira)
            if 'Data de apuração' in txt or 'Data de apura' in txt:
                # Extrai 'Data de apuração: dd/mm/yyyy'
                m = re.search(r'(\d{2}/\d{2}/\d{4})', txt)
                if m:
                    data_apuracao = m.group(1)
            if 'TOTAL GERAL DA AÇÃO' in txt.upper() or 'TOTAL GERAL DA ACAO' in txt.upper():
                # Valor está na coluna 8 (H), índice 7
                if len(row) > 7 and row[7] is not None:
                    try:
                        total_geral = float(row[7])
                    except (TypeError, ValueError):
                        pass
            elif 'SUBTOTAL' in txt.upper():
                if len(row) > 7 and row[7] is not None:
                    try:
                        subtotal_dobrado = float(row[7])
                    except (TypeError, ValueError):
                        pass
            elif 'DANO MORAL' in txt.upper():
                if len(row) > 7 and row[7] is not None:
                    try:
                        dano_moral = float(row[7])
                    except (TypeError, ValueError):
                        pass
        if total_geral is None:
            return None
        return {
            'total_geral': total_geral,
            'subtotal_dobrado': subtotal_dobrado,
            'dano_moral': dano_moral,
            'data_apuracao': data_apuracao,
        }
    except Exception:
        return None


# Nome canônico do arquivo Excel gerado pela kit-juridico (sem cliente/banco
# no nome — uma pasta de ação = um cálculo). A inicial procura por esse nome
# para reusar o cálculo.
NOME_CANONICO_EXCEL_KIT = 'CALCULO_INDEBITO.xlsx'


def localizar_excel_indebito(pasta_acao: str) -> Optional[str]:
    """Procura um Excel de cálculo já gerado na pasta_acao.

    Tenta nesta ordem:
      1. CALCULO_INDEBITO.xlsx (nome canônico kit-juridico)
      2. Qualquer CALCULO_*.xlsx (compatibilidade com Excels gerados pela
         inicial em sessões anteriores)
    """
    if not pasta_acao or not os.path.isdir(pasta_acao):
        return None
    # 1. Nome canônico
    canonico = os.path.join(pasta_acao, NOME_CANONICO_EXCEL_KIT)
    if os.path.exists(canonico):
        return canonico
    # 2. Qualquer CALCULO_*.xlsx
    for f in os.listdir(pasta_acao):
        if f.upper().startswith('CALCULO_') and f.lower().endswith('.xlsx'):
            return os.path.join(pasta_acao, f)
    return None


if __name__ == '__main__':
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    # Teste rápido
    contratos_teste = [
        {
            'numero': '622902175',
            'banco_nome': 'BANCO ITAU CONSIGNADO SA',
            'valor_parcela': 'R$49,50',
            'qtd_parcelas': 84,
            'competencia_inicio_str': '08/2020',
            'competencia_fim_str': '07/2027',
            'situacao': 'Ativo',
        },
    ]
    out = gerar_excel_indebito(contratos_teste, 'TESTE FULANO',
                                 r'C:\Users\gabri\OneDrive\Área de Trabalho\CLAUDE\_teste_calculo.xlsx')
    print(f'Excel gerado: {out}')
