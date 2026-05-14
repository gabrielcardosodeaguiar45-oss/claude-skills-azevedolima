#!/usr/bin/env python3
"""
Gera a Planilha de Pendências em formato Excel (.xlsx).

Uso:
    python gerar_pendencias.py <output_path> <pendencias_json>

O JSON de entrada deve ter o formato:
[
    {
        "categoria": "Comprovante de residência",
        "pendencia": "Possível falsificação",
        "observacao": "Fontes inconsistentes detectadas",
        "status": "Pendente"
    }
]
"""

import sys
import os
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    raise ImportError(
        f"Dependência ausente: {e}. "
        f"Instale via: pip install -r requirements.txt"
    ) from e


def create_pendencias_xlsx(output_path, pendencias):
    """
    Cria planilha Excel formatada com as pendências encontradas.

    Args:
        output_path: Caminho para salvar o arquivo .xlsx
        pendencias: Lista de dicts com keys: categoria, pendencia, observacao, status
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendências"

    # Estilos
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Calibri', size=11)
    cell_alignment = Alignment(vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    status_pending_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    status_resolved_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

    # Cabeçalhos
    headers = ['Nº', 'Categoria', 'Pendência', 'Observação', 'Status']
    col_widths = [6, 30, 40, 50, 15]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Dados
    for row_idx, pend in enumerate(pendencias, 2):
        num = row_idx - 1

        values = [
            num,
            pend.get('categoria', ''),
            pend.get('pendencia', ''),
            pend.get('observacao', ''),
            pend.get('status', 'Pendente')
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Colorir status
            if col_idx == 5:
                if value == 'Pendente':
                    cell.fill = status_pending_fill
                elif value == 'Resolvido':
                    cell.fill = status_resolved_fill

    # Congelar painel do cabeçalho
    ws.freeze_panes = 'A2'

    # Auto-filtro
    if pendencias:
        ws.auto_filter.ref = f"A1:E{len(pendencias) + 1}"

    # Título
    ws.sheet_properties.tabColor = "2F5496"

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 3:
        print("Uso: python gerar_pendencias.py <output.xlsx> <pendencias.json>")
        sys.exit(1)

    output_path = sys.argv[1]
    json_path = sys.argv[2]

    with open(json_path, 'r', encoding='utf-8') as f:
        pendencias = json.load(f)

    create_pendencias_xlsx(output_path, pendencias)
    print(f"Planilha criada: {output_path}")
    print(f"Total de pendências: {len(pendencias)}")


if __name__ == "__main__":
    main()
