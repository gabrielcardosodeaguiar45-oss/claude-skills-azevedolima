#!/usr/bin/env python3
"""
Gera relatório consolidado de processamento em lote de kits jurídicos.

Uso:
    python gerar_relatorio_lote.py <output_xlsx> <resumos_json>

O JSON de entrada deve ter o formato:
{
    "kits": [
        {
            "cliente": "Maria da Silva Santos",
            "acoes": ["RMC"],
            "pendencias": [
                {"categoria": "...", "pendencia": "...", "observacao": "...", "status": "Pendente"}
            ],
            "status": "ok",
            "erro": null
        },
        {
            "cliente": "João Oliveira",
            "acoes": ["RMC", "RCC"],
            "pendencias": [],
            "status": "ok",
            "erro": null
        },
        {
            "cliente": "Pasta com erro",
            "acoes": [],
            "pendencias": [],
            "status": "erro",
            "erro": "KIT assinado não encontrado"
        }
    ]
}
"""

import sys
import os
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    raise ImportError(
        f"Dependência ausente: {e}. "
        f"Instale via: pip install -r requirements.txt"
    ) from e


def create_relatorio_consolidado(output_path, data):
    """
    Cria planilha Excel com 3 abas: Resumo, Pendências, Erros.
    """
    wb = Workbook()

    # Estilos comuns
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Calibri', size=11)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ok_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    pending_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    error_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')

    def write_header(ws, headers, widths):
        for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
            ws.column_dimensions[col_letter].width = width

    def write_cell(ws, row, col, value, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill
        return cell

    kits = data.get('kits', [])

    # === ABA 1: RESUMO ===
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.sheet_properties.tabColor = "2F5496"

    headers = ['Nº', 'Cliente', 'Ações Identificadas', 'Total Pendências', 'Status']
    widths = [6, 35, 30, 18, 18]
    write_header(ws_resumo, headers, widths)

    for i, kit in enumerate(kits, 1):
        row = i + 1
        status = "Erro" if kit.get('status') == 'erro' else ("Com pendências" if kit.get('pendencias') else "OK")
        fill = error_fill if status == "Erro" else (pending_fill if status == "Com pendências" else ok_fill)

        write_cell(ws_resumo, row, 1, i)
        write_cell(ws_resumo, row, 2, kit.get('cliente', ''))
        write_cell(ws_resumo, row, 3, ', '.join(kit.get('acoes', [])))
        write_cell(ws_resumo, row, 4, len(kit.get('pendencias', [])))
        write_cell(ws_resumo, row, 5, status, fill=fill)

    ws_resumo.freeze_panes = 'A2'
    if kits:
        ws_resumo.auto_filter.ref = f"A1:E{len(kits) + 1}"

    # === ABA 2: PENDÊNCIAS ===
    ws_pend = wb.create_sheet("Pendências")
    ws_pend.sheet_properties.tabColor = "FFC000"

    headers = ['Nº', 'Cliente', 'Categoria', 'Pendência', 'Observação', 'Status']
    widths = [6, 30, 25, 35, 45, 15]
    write_header(ws_pend, headers, widths)

    row = 2
    for kit in kits:
        for pend in kit.get('pendencias', []):
            write_cell(ws_pend, row, 1, row - 1)
            write_cell(ws_pend, row, 2, kit.get('cliente', ''))
            write_cell(ws_pend, row, 3, pend.get('categoria', ''))
            write_cell(ws_pend, row, 4, pend.get('pendencia', ''))
            write_cell(ws_pend, row, 5, pend.get('observacao', ''))
            write_cell(ws_pend, row, 6, pend.get('status', 'Pendente'), fill=pending_fill)
            row += 1

    ws_pend.freeze_panes = 'A2'
    if row > 2:
        ws_pend.auto_filter.ref = f"A1:F{row - 1}"

    # === ABA 3: ERROS ===
    ws_erros = wb.create_sheet("Erros")
    ws_erros.sheet_properties.tabColor = "FF0000"

    headers = ['Nº', 'Cliente', 'Motivo do Erro']
    widths = [6, 35, 60]
    write_header(ws_erros, headers, widths)

    row = 2
    for kit in kits:
        if kit.get('status') == 'erro':
            write_cell(ws_erros, row, 1, row - 1)
            write_cell(ws_erros, row, 2, kit.get('cliente', ''))
            write_cell(ws_erros, row, 3, kit.get('erro', 'Erro desconhecido'), fill=error_fill)
            row += 1

    ws_erros.freeze_panes = 'A2'

    # Salvar
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 3:
        print("Uso: python gerar_relatorio_lote.py <output.xlsx> <resumos.json>")
        sys.exit(1)

    output_path = sys.argv[1]
    json_path = sys.argv[2]

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    create_relatorio_consolidado(output_path, data)

    kits = data.get('kits', [])
    total = len(kits)
    ok = sum(1 for k in kits if k.get('status') != 'erro' and not k.get('pendencias'))
    com_pend = sum(1 for k in kits if k.get('status') != 'erro' and k.get('pendencias'))
    erros = sum(1 for k in kits if k.get('status') == 'erro')

    print(f"Relatório gerado: {output_path}")
    print(f"Total: {total} kits | OK: {ok} | Com pendências: {com_pend} | Erros: {erros}")


if __name__ == "__main__":
    main()
