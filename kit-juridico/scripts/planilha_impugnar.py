"""
Gera/lê a planilha _contratos_a_impugnar.xlsx — formato de revisão humana
do `contratos_impugnar_ids` do `_estado_cliente.json`.

Formato:
  | pasta_acao | id | contrato | banco | tipo | situacao | motivo | flags | impugnar (S/N) |

Após o advogado revisar, a coluna `impugnar` é lida e atualiza
`pastas_acao[].contratos_impugnar_ids` no JSON.
"""
import os


def gerar_planilha(linhas: list, output_path: str) -> str | None:
    """Cria xlsx com as linhas. Retorna path ou None se sem dados."""
    if not linhas:
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system('pip install openpyxl --break-system-packages -q')
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos a Impugnar'
    headers = ['pasta_acao', 'id', 'contrato', 'banco', 'tipo',
               'situacao', 'motivo', 'flags', 'impugnar (S/N)']
    ws.append(headers)
    # Estilo header
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='B4C7E7')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    for ln in linhas:
        ws.append([
            ln.get('pasta_acao', ''),
            ln.get('id', ''),
            str(ln.get('contrato') or ''),
            ln.get('banco', ''),
            ln.get('tipo', ''),
            ln.get('situacao', ''),
            ln.get('motivo', ''),
            ln.get('flags', ''),
            ln.get('impugnar', 'S'),
        ])

    # Larguras
    larguras = [40, 6, 18, 12, 12, 12, 22, 24, 14]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Linha de instruções no topo (acima do header não dá; coloca em sheet separada)
    ws_help = wb.create_sheet('LEIA-ME')
    ws_help.append(['Instruções'])
    ws_help.append([])
    ws_help.append(['1. Esta planilha contém a sugestão automática da kit-juridico de'])
    ws_help.append(['   quais contratos do HISCON serão objeto de ação por pasta.'])
    ws_help.append(['2. Revise a coluna "impugnar (S/N)" linha a linha.'])
    ws_help.append(['3. Mude para "N" o que NÃO deve impugnar (ex: prescrito, conta de'])
    ws_help.append(['   terceiros, contrato legítimo do cliente).'])
    ws_help.append(['4. Salve o arquivo e rode kit-juridico --modo=revisar para o JSON'])
    ws_help.append(['   ser atualizado.'])
    ws_help.append([])
    ws_help.append(['Flags:'])
    ws_help.append(['  revisar_prescricao = contrato encerrado/excluído isolado.'])
    ws_help.append(['     Vale impugnar se o último desconto foi há menos de 5 anos.'])
    ws_help.append(['  operacao_ponte = banco_codigo_inss difere do banco_pagador.'])
    ws_help.append(['     Pode ser fraude do tipo "lançamento morto".'])
    ws_help.column_dimensions['A'].width = 80

    wb.save(output_path)
    return output_path


def ler_planilha(planilha_path: str) -> dict:
    """Lê uma planilha já revisada e retorna dict {pasta_acao: [ids_a_impugnar]}.

    Usado pelo modo --revisar.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        os.system('pip install openpyxl --break-system-packages -q')
        from openpyxl import load_workbook

    if not os.path.exists(planilha_path):
        return {}
    wb = load_workbook(planilha_path, data_only=True)
    if 'Contratos a Impugnar' in wb.sheetnames:
        ws = wb['Contratos a Impugnar']
    else:
        ws = wb.active

    out: dict[str, list[str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row or not row[0]:
            continue
        pasta = row[0]
        cid = row[1]
        impugnar = (str(row[8] or '').strip().upper())
        if impugnar == 'S' and cid:
            out.setdefault(pasta, []).append(cid)
    return out
