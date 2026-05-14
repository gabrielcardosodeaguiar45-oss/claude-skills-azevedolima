"""Auditor + completador automático para teses TARIFAS no Bradesco.

Quando a tabela XLSX do NotebookLM está incompleta (ignorou VR.PARCIAL,
EMISSÃO EXTRATO, etc.), esta rotina:

1. Procura extrato digital (text-layer) na pasta do cliente ou no `0. Kit/`
2. Faz parsing posicional preciso (parser_extrato_posicional.parsear_extrato_digital)
3. Filtra TODOS os lançamentos com 'TARIFA' na descrição
4. Compara com a tabela do NotebookLM
5. Se a tabela está incompleta, gera planilha v2 substituta na pasta do cliente
6. Retorna lançamentos completos + flags + relatório

Use no `_run_*_tarifas.py` para garantir que TODA tarifa entre na inicial.
"""
import os
import sys
from typing import Dict, List, Optional


def auditar_e_completar_tarifas(
    pasta_cliente: str,
    tabela_xlsx_path: Optional[str] = None,
    cliente_nome: str = '',
    conta_label: str = '',
    procuracao_label: str = 'TARIFA BANCÁRIA - CESTA B.EXPRESSO',
    gerar_planilha_v2: bool = True,
) -> Dict:
    """Cruza tabela do NotebookLM com extrato digital e gera planilha v2 completa.

    Args:
        pasta_cliente: pasta raiz do cliente (onde está '0. Kit/' ou similar)
        tabela_xlsx_path: planilha original do NotebookLM (opcional, para comparação)
        cliente_nome: ex 'CELIA RODRIGUES DA SILVA'
        conta_label: ex 'Agência: 3706 | Conta: 16649-9'
        procuracao_label: rubrica da procuração (default 'TARIFA BANCÁRIA - CESTA B.EXPRESSO')
        gerar_planilha_v2: se True, gera nova planilha XLSX na pasta do cliente

    Returns:
        {
            'extrato_digital_encontrado': bool,
            'extrato_digital_path': str | None,
            'lancamentos': [{'data': str, 'descricao': str, 'valor': float}, ...],
            'lancamentos_xlsx_original': [...],   # se tabela_xlsx_path foi fornecida
            'qtd_extrato': int,
            'qtd_xlsx_original': int,
            'soma_extrato': float,
            'soma_xlsx_original': float,
            'planilha_v2_path': str | None,
            'severidade': 'OK' | 'INCOMPLETO' | 'CRITICO',
            'recomendacao': str,
        }
    """
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from parser_extrato_posicional import (
        parsear_extrato_digital, filtrar_por_palavra_chave,
        encontrar_extratos_digitais, parsear_multiplos_extratos,
        classificar_tarifa,
    )
    from helpers_xlsx import gerar_xlsx_tarifas

    resultado = {
        'extrato_digital_encontrado': False,
        'extrato_digital_path': None,
        'extrato_digital_paths': [],
        'lancamentos': [],
        'lancamentos_xlsx_original': [],
        'qtd_extrato': 0,
        'qtd_xlsx_original': 0,
        'soma_extrato': 0.0,
        'soma_xlsx_original': 0.0,
        'planilha_v2_path': None,
        'severidade': 'OK',
        'recomendacao': '',
    }

    # 1. Localizar TODOS os extratos digitais (cliente pode ter fragmentado em vários PDFs)
    extrato_paths = encontrar_extratos_digitais(pasta_cliente)
    if not extrato_paths:
        resultado['severidade'] = 'CRITICO'
        resultado['recomendacao'] = (
            'Extrato digital com text-layer não encontrado na pasta do cliente '
            'nem no KIT. Necessário extrato digital baixado direto do app Bradesco. '
            'Sem ele, só OCR (impreciso). Solicitar ao cliente.'
        )
        return resultado
    resultado['extrato_digital_encontrado'] = True
    resultado['extrato_digital_paths'] = extrato_paths
    resultado['extrato_digital_path'] = extrato_paths[0]  # primeiro como referência

    # 2. Parsing posicional de TODOS os extratos (com de-dup)
    eventos = parsear_multiplos_extratos(extrato_paths)
    if not eventos:
        resultado['severidade'] = 'CRITICO'
        resultado['recomendacao'] = (
            f'Extratos digitais em {pasta_cliente} sem text-layer útil. '
            'Verificar arquivos ou usar OCR.'
        )
        return resultado

    # 3. Filtrar TUDO com 'TARIFA'
    tarifas = filtrar_por_palavra_chave(eventos, 'TARIFA')
    resultado['lancamentos'] = tarifas
    resultado['qtd_extrato'] = len(tarifas)
    resultado['soma_extrato'] = round(sum(t['valor'] for t in tarifas), 2)

    # 4. Comparar com XLSX original (se fornecido)
    if tabela_xlsx_path and os.path.exists(tabela_xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(tabela_xlsx_path, data_only=True)
            lancs_xlsx = []
            # Procura abas relevantes
            for sheet_name in wb.sheetnames:
                if 'TARIFA' not in sheet_name.upper() and 'CESTA' not in sheet_name.upper():
                    continue
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 6 or not row or not row[0]:
                        continue
                    # Header DATA/NOME/VALOR
                    if isinstance(row[0], str):
                        if 'TOTAL' in row[0].upper() or 'QUANTIDADE' in row[0].upper():
                            continue
                    try:
                        v = float(row[2]) if row[2] else 0
                        if v > 0:
                            lancs_xlsx.append({
                                'data': row[0].strftime('%d/%m/%Y') if hasattr(row[0], 'strftime') else str(row[0]),
                                'descricao': row[1] or '',
                                'valor': v,
                            })
                    except (ValueError, TypeError, AttributeError):
                        pass
            resultado['lancamentos_xlsx_original'] = lancs_xlsx
            resultado['qtd_xlsx_original'] = len(lancs_xlsx)
            resultado['soma_xlsx_original'] = round(sum(l['valor'] for l in lancs_xlsx), 2)
        except Exception as e:
            resultado['recomendacao'] += f'Erro ao ler XLSX original: {e}. '

    # 5. Decidir severidade
    diff_qtd = resultado['qtd_extrato'] - resultado['qtd_xlsx_original']
    diff_soma = resultado['soma_extrato'] - resultado['soma_xlsx_original']
    if resultado['qtd_xlsx_original'] == 0:
        resultado['severidade'] = 'INCOMPLETO'
        resultado['recomendacao'] = (
            f'Tabela do NotebookLM ausente ou vazia. Usar dados do extrato digital '
            f'({resultado["qtd_extrato"]} lançamentos / R$ {resultado["soma_extrato"]:.2f}).'
        )
    elif diff_qtd > 5 or abs(diff_soma) > 50:
        resultado['severidade'] = 'INCOMPLETO'
        resultado['recomendacao'] = (
            f'Tabela do NotebookLM tem {resultado["qtd_xlsx_original"]} lançamentos / '
            f'R$ {resultado["soma_xlsx_original"]:.2f}, mas extrato digital tem '
            f'{resultado["qtd_extrato"]} / R$ {resultado["soma_extrato"]:.2f}. '
            f'Diferença: +{diff_qtd} lançamentos / R$ {diff_soma:+.2f}. '
            'A skill usará os dados do extrato direto.'
        )
    elif diff_qtd > 0:
        resultado['severidade'] = 'OK'
        resultado['recomendacao'] = (
            f'Tabela e extrato batem aproximadamente. Pequena diferença '
            f'(+{diff_qtd} lançamentos) absorvida.'
        )

    # 6. Gerar planilha v2 (substitui a do NotebookLM)
    if gerar_planilha_v2 and resultado['severidade'] in ('INCOMPLETO', 'CRITICO'):
        # Nome v2 ao lado da original — não duplica "- v2" se já existir
        if tabela_xlsx_path:
            base = tabela_xlsx_path.rsplit('.', 1)[0]
            if base.endswith(' - v2'):
                xlsx_v2 = base + '.xlsx'  # já é v2; sobrescreve
            else:
                xlsx_v2 = base + ' - v2.xlsx'
        else:
            # fallback: na raiz da pasta_cliente
            xlsx_v2 = os.path.join(
                pasta_cliente,
                f'Tabela de Descontos por Procuracao - {cliente_nome} - v2.xlsx',
            )
        try:
            gerar_xlsx_tarifas(
                lancamentos=tarifas,
                output_path=xlsx_v2,
                cliente_nome=cliente_nome,
                conta_label=conta_label,
                procuracao_label=procuracao_label,
                classificador=classificar_tarifa,
            )
            resultado['planilha_v2_path'] = xlsx_v2
        except Exception as e:
            resultado['recomendacao'] += f'Erro ao gerar planilha v2: {e}. '

    return resultado


def lancamentos_para_tese(lancamentos: List[Dict]) -> List[tuple]:
    """Converte lista de dicts em formato esperado pela `tese['lancamentos']`
    de `montar_dados_padrao`: lista de tuplas (data_str, valor_float).
    """
    return [(ev['data'], ev['valor']) for ev in lancamentos]
