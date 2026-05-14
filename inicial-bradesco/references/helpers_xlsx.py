"""Helpers de geração de XLSX no formato esperado pelo escritório
(equivalente ao que o NotebookLM produz, mas com TODOS os lançamentos).

Use quando a tabela XLSX original do NotebookLM estiver incompleta e a
skill detectar precisa substituí-la com base no parsing direto do extrato.
"""
import os
from typing import List, Dict, Optional
from datetime import datetime


def gerar_xlsx_tarifas(
    lancamentos: List[Dict],
    output_path: str,
    cliente_nome: str,
    banco_label: str = 'Banco Bradesco S/A',
    conta_label: str = '',
    procuracao_label: str = '',
    classificador=None,
) -> str:
    """Gera planilha XLSX no formato:
        - Aba RESUMO (totais por categoria)
        - Aba TODAS (todos os lançamentos consolidados)
        - 1 aba por categoria detectada (via classificador)

    Cada aba tem cabeçalho institucional + tabela DATA/NOME/VALOR + linha TOTAL.

    Args:
        lancamentos: lista de dicts {data: 'DD/MM/YYYY', descricao: str, valor: float}
        output_path: caminho .xlsx de destino
        cliente_nome: nome completo (para cabeçalho)
        banco_label: ex 'Banco Bradesco S/A'
        conta_label: ex 'Agência: 3706 | Conta: 16649-9'
        procuracao_label: ex 'TARIFA BANCÁRIA - CESTA B.EXPRESSO'
        classificador: função(descricao) → categoria (string). Se None,
                        tudo vai pra 1 categoria 'GERAL'.

    Retorna o path salvo.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system('pip install openpyxl --break-system-packages -q')
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not lancamentos:
        return ''

    # Adicionar categoria + data_dt
    if classificador is None:
        for ev in lancamentos:
            ev['_categoria'] = 'GERAL'
    else:
        for ev in lancamentos:
            ev['_categoria'] = classificador(ev.get('descricao', ''))
    for ev in lancamentos:
        if isinstance(ev.get('data'), str):
            ev['_data_dt'] = datetime.strptime(ev['data'], '%d/%m/%Y')
        else:
            ev['_data_dt'] = ev.get('data')

    lancamentos.sort(key=lambda x: x['_data_dt'])

    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    bold_white = Font(bold=True, color='FFFFFF')
    fill_header = PatternFill('solid', fgColor='305496')
    fill_subheader = PatternFill('solid', fgColor='D9E1F2')
    center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def _criar_aba(nome_aba: str, titulo: str, lst: List[Dict]):
        ws = wb.create_sheet(nome_aba)
        ws.append([titulo, None, None])
        ws.merge_cells('A1:C1')
        c1 = ws['A1']; c1.font = Font(bold=True, size=12); c1.alignment = center

        ws.append([f'Cliente: {cliente_nome}', None, None]); ws.merge_cells('A2:C2')
        ws['A2'].font = bold

        ws.append([f'{banco_label} | {conta_label}', None, None]); ws.merge_cells('A3:C3')
        ws['A3'].font = bold

        ws.append([f'Procuração: {procuracao_label}', None, None]); ws.merge_cells('A4:C4')
        ws['A4'].font = bold

        ws.append([None, None, None])  # vazia

        ws.append(['DATA', 'NOME', 'VALOR DO DESCONTO (R$)'])
        for col in range(1, 4):
            c = ws.cell(row=6, column=col)
            c.font = bold_white; c.fill = fill_header
            c.alignment = center; c.border = border_thin

        for ev in lst:
            ws.append([ev['_data_dt'].date(), ev.get('descricao', ''), ev.get('valor', 0)])

        for r in range(7, 7 + len(lst)):
            for col in range(1, 4):
                c = ws.cell(row=r, column=col)
                c.border = border_thin
                if col == 1: c.number_format = 'DD/MM/YYYY'
                if col == 3: c.number_format = '#,##0.00'

        # Linha TOTAL
        fim = 7 + len(lst)
        ws.cell(row=fim, column=1, value='TOTAL').font = bold
        ws.cell(row=fim, column=2, value=f'{len(lst)} lançamentos').font = bold
        soma = sum(e.get('valor', 0) for e in lst)
        c_soma = ws.cell(row=fim, column=3, value=soma)
        c_soma.font = bold; c_soma.number_format = '#,##0.00'
        for col in range(1, 4):
            ws.cell(row=fim, column=col).fill = fill_subheader

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 22

    # Aba TODAS
    _criar_aba('TARIFAS - TODAS', f'TODAS AS TARIFAS — {cliente_nome.upper()}', lancamentos)

    # Abas por categoria
    categorias = {}
    for ev in lancamentos:
        categorias.setdefault(ev['_categoria'], []).append(ev)

    for cat, lst in categorias.items():
        nome_aba = cat[:31]
        # Se nome aba conflitar com 'TARIFAS - TODAS', adiciona sufixo
        if nome_aba == 'TARIFAS - TODAS':
            nome_aba = nome_aba + ' (cat)'
        _criar_aba(nome_aba, cat, lst)

    # Aba RESUMO
    ws = wb.create_sheet('RESUMO')
    ws.append([f'RESUMO — {cliente_nome}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(['Categoria', 'Lançamentos', 'Soma'])
    for c in ws[3]: c.font = bold
    for cat, lst in categorias.items():
        ws.append([cat, len(lst), round(sum(e.get('valor', 0) for e in lst), 2)])
    soma_total = sum(e.get('valor', 0) for e in lancamentos)
    ws.append(['TOTAL', len(lancamentos), round(soma_total, 2)])
    ultima = ws.max_row
    for c in ws[ultima]: c.font = bold
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18

    wb.move_sheet('RESUMO', offset=-len(wb.sheetnames) + 1)

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path
